"""
Exportacion de datasets a Excel (.xlsx) y CSV con formato consistente.

Provee:
- to_xlsx_bytes(columns, rows, sheet_name) → bytes del Excel
- to_csv_string(columns, rows) → CSV UTF-8 con BOM (Excel-friendly en AR)
- export_catalog() → lista de exportes pre-armados para el centro de exports
"""
from __future__ import annotations

import csv
import io
import logging
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.utils.tz import now_ar

log = logging.getLogger("unidata.exports")


# Identidad visual Unistore para los Excels
HEADER_FILL = PatternFill(start_color="7A3EAE", end_color="7A3EAE", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
CELL_FONT = Font(name="Calibri", size=10)


def to_xlsx_bytes(
    columns: list[str],
    rows: list[list[Any]],
    sheet_name: str = "Datos",
    title: str | None = None,
) -> bytes:
    """Genera un Excel xlsx con encabezado violeta + auto-width columnas.
    Devuelve bytes listos para descarga."""
    wb = Workbook()
    ws = wb.active
    ws.title = (sheet_name[:30] or "Datos")  # Excel limite 31 chars

    start_row = 1
    if title:
        ws.cell(row=1, column=1, value=title).font = Font(name="Calibri", bold=True, size=14, color="1F1235")
        ws.cell(row=2, column=1, value=f"Generado: {now_ar().strftime('%Y-%m-%d %H:%M')} (AR)").font = Font(name="Calibri", size=9, color="6B7280")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(columns)))
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(1, len(columns)))
        start_row = 4

    # Header
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=str(col_name))
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[start_row].height = 22

    # Data
    for row_idx, row in enumerate(rows, start=start_row + 1):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = CELL_FONT
            # Numeros como numeros, otros como texto
            if isinstance(value, (int, float)) and not isinstance(value, bool):
                cell.number_format = "#,##0.00" if isinstance(value, float) else "#,##0"

    # Auto-width (con cap a 60)
    for col_idx, col_name in enumerate(columns, start=1):
        max_len = len(str(col_name))
        for row in rows:
            if col_idx <= len(row):
                v = row[col_idx - 1]
                if v is not None:
                    max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(60, max(10, max_len + 2))

    # Freeze header
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def to_csv_string(columns: list[str], rows: list[list[Any]]) -> str:
    """CSV UTF-8 con BOM para que Excel en Windows reconozca acentos."""
    buf = io.StringIO()
    buf.write("﻿")  # BOM
    writer = csv.writer(buf, delimiter=",", quotechar='"', quoting=csv.QUOTE_MINIMAL)
    writer.writerow(columns)
    for row in rows:
        writer.writerow(["" if v is None else v for v in row])
    return buf.getvalue()


# ============================================================
# Catalogo de exportes pre-armados para el centro
# ============================================================

EXPORT_CATALOG = [
    {
        "id": "vip_marketing",
        "label": "Clientes VIP — Campañas marketing",
        "description": "Lista de clientes VIP Unistore con email, teléfono, tier y última compra. Para envío de campañas a Gold/Silver/Bronze.",
        "team": "Marketing",
        "icon": "crown",
        "endpoint": "/api/exports/vip-marketing",
        "format": "xlsx",
        "fields": ["customer_id", "cliente", "email", "telefono", "tier", "lifetime", "ordenes", "ultima_compra", "recency_dias", "provincia"],
    },
    {
        "id": "customers_at_risk",
        "label": "Clientes Posible Churn — Retención",
        "description": "Clientes recurrentes que excedieron 1.5x su gap promedio. Accionables para reactivación.",
        "team": "Customer Success",
        "icon": "alert",
        "endpoint": "/api/exports/customers-at-risk",
        "format": "xlsx",
        "fields": ["customer_id", "cliente", "email", "telefono", "lifetime", "dias_desde_ultima", "avg_gap_days"],
    },
    {
        "id": "dropshippers_active",
        "label": "Dropshippers activos — Comercial",
        "description": "Operadores Unidrop con suscripción activa y al menos una venta en los últimos 30 días.",
        "team": "Comercial",
        "icon": "users",
        "endpoint": "/api/exports/dropshippers-active",
        "format": "xlsx",
        "fields": ["user_id", "nombre", "email", "telefono", "plan", "gmv_30d", "deuda_pendiente"],
    },
    {
        "id": "stuck_orders",
        "label": "Pedidos atascados — Logística",
        "description": "Órdenes pagadas sin fulfillment hace más de 5 días. Para revisar urgente.",
        "team": "Logística",
        "icon": "truck",
        "endpoint": "/api/exports/stuck-orders",
        "format": "xlsx",
        "fields": ["id", "numero", "fecha", "cliente", "total", "provincia"],
    },
    {
        "id": "subs_expiring",
        "label": "Suscripciones MELI por vencer — Cobranza",
        "description": "Dropshippers con suscripción que vence en próximos 7 días. Para campaña de renovación.",
        "team": "Cobranza",
        "icon": "calendar",
        "endpoint": "/api/exports/subs-expiring",
        "format": "xlsx",
        "fields": ["user_id", "nombre", "email", "telefono", "plan", "dias_al_vencimiento"],
    },
    {
        "id": "stock_critico",
        "label": "Stock crítico — Compras",
        "description": "SKUs con menos de 5 unidades en depósito. Para reposición urgente.",
        "team": "Compras",
        "icon": "package",
        "endpoint": "/api/exports/stock-critico",
        "format": "xlsx",
        "fields": ["sku", "nombre", "stock_actual", "ventas_30d"],
    },
    {
        "id": "top_skus_30d",
        "label": "Top SKUs últimos 30 días — Producto",
        "description": "Ranking de productos por revenue cross-canal (TN + ML). Para análisis de oferta.",
        "team": "Producto",
        "icon": "shopping-bag",
        "endpoint": "/api/exports/top-skus-30d",
        "format": "xlsx",
        "fields": ["sku", "nombre", "ean", "revenue", "unidades", "ordenes"],
    },
    {
        "id": "devoluciones_abiertas",
        "label": "Devoluciones abiertas — Customer Service",
        "description": "Casos de devolución sin resolver. Para asignar y seguir.",
        "team": "Customer Service",
        "icon": "rotate",
        "endpoint": "/api/exports/devoluciones-abiertas",
        "format": "xlsx",
        "fields": ["id", "fecha", "cliente_email", "modelo", "monto", "estado"],
    },
]


def get_catalog() -> dict:
    """Devuelve el catalogo agrupado por team."""
    by_team: dict[str, list[dict]] = {}
    for exp in EXPORT_CATALOG:
        team = exp["team"]
        by_team.setdefault(team, []).append(exp)
    return {
        "items": EXPORT_CATALOG,
        "by_team": by_team,
        "teams": sorted(by_team.keys()),
        "total": len(EXPORT_CATALOG),
        "generated_at": now_ar().isoformat(),
    }
