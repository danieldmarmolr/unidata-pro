"""
Parsers e importadores Excel del modulo Flujo de Fondos.

3 plantillas soportadas:
- erogaciones: fecha_pago, descripcion, monto, empresa, banco, proveedor (opt),
               estado (opt), categoria (opt), notas (opt)
- ingresos_puntuales: fecha, descripcion, monto, empresa, banco (opt),
                       categoria (opt), notas (opt)
- facturacion: fecha, unidad_negocio, empresa (opt), monto, es_real (opt),
               es_evento_puntual (opt)

Patron: el frontend sube .xlsx, backend lo parsea, devuelve preview con
ok/warnings/errores. Frontend muestra preview y al confirmar llama aplicar
que inserta en bulk con detección de duplicados.
"""
from __future__ import annotations

import io
import logging
from datetime import date, datetime
from typing import Any

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.utils import get_column_letter

from app.db.local_persistence import get_conn

log = logging.getLogger("flujo_fondos.importadores")


def _parse_date(v: Any) -> str | None:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def _parse_number(v: Any) -> float | None:
    if v is None or v == "":
        return None
    try:
        if isinstance(v, str):
            v = v.replace(".", "").replace(",", ".") if "," in v and v.count(",") == 1 else v.replace(",", "")
        return float(v)
    except (ValueError, TypeError):
        return None


def _load_maestros() -> dict:
    """Carga empresas, bancos, proveedores, unidades_negocio en dicts por nombre normalizado."""
    out: dict[str, dict[str, int]] = {"empresas": {}, "bancos": {}, "proveedores": {}, "unidades_negocio": {}}
    with get_conn() as c, c.cursor() as cur:
        cur.execute('SELECT id, nombre FROM public."empresas"')
        for r in cur.fetchall():
            out["empresas"][r["nombre"].strip().lower()] = int(r["id"])
        cur.execute('SELECT id, nombre FROM public."bancos_medios_pago"')
        for r in cur.fetchall():
            out["bancos"][r["nombre"].strip().lower()] = int(r["id"])
        cur.execute('SELECT id, nombre FROM public."proveedores"')
        for r in cur.fetchall():
            out["proveedores"][r["nombre"].strip().lower()] = int(r["id"])
        cur.execute('SELECT id, nombre FROM public."unidades_negocio"')
        for r in cur.fetchall():
            out["unidades_negocio"][r["nombre"].strip().lower()] = int(r["id"])
    return out


# ============================================================
# Plantilla erogaciones
# ============================================================

PLANTILLA_EROGACIONES_HEADERS = [
    "fecha_pago", "descripcion", "monto", "empresa", "banco",
    "proveedor", "estado", "categoria", "notas",
]


def generar_plantilla_erogaciones() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "EROGACIONES"
    for i, h in enumerate(PLANTILLA_EROGACIONES_HEADERS, 1):
        ws.cell(row=1, column=i, value=h)
        ws.column_dimensions[get_column_letter(i)].width = 18
    # ejemplos
    ws.cell(row=2, column=1, value="2026-05-25")
    ws.cell(row=2, column=2, value="Pago alquiler oficina")
    ws.cell(row=2, column=3, value=350000)
    ws.cell(row=2, column=4, value="FOX ELECTRONICS")
    ws.cell(row=2, column=5, value="SUPERVIELLE")
    ws.cell(row=2, column=6, value="Inmobiliaria SA")
    ws.cell(row=2, column=7, value="pendiente")
    ws.cell(row=2, column=8, value="alquiler")
    ws.cell(row=2, column=9, value="contrato firmado oct/24")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parsear_erogaciones(file_bytes: bytes) -> dict:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    if ws is None:
        return {"items": [], "errors": ["Excel vacio"], "warnings": []}
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return {"items": [], "errors": ["Solo hay headers"], "warnings": []}

    maestros = _load_maestros()
    items: list[dict] = []
    errors: list[str] = []
    warnings: list[str] = []
    empresas_faltantes: set[str] = set()
    bancos_faltantes: set[str] = set()

    headers = [str(h).strip().lower() if h else "" for h in rows[0]]
    idx = {h: i for i, h in enumerate(headers)}

    for ri, row in enumerate(rows[1:], 2):
        if not any(c is not None and c != "" for c in row):
            continue
        fecha = _parse_date(row[idx.get("fecha_pago", 0)] if "fecha_pago" in idx else None)
        descripcion = str(row[idx.get("descripcion", 1)] or "").strip() if "descripcion" in idx else ""
        monto = _parse_number(row[idx.get("monto", 2)] if "monto" in idx else None)
        empresa_str = str(row[idx.get("empresa", 3)] or "").strip() if "empresa" in idx else ""
        banco_str = str(row[idx.get("banco", 4)] or "").strip() if "banco" in idx else ""
        proveedor_str = str(row[idx.get("proveedor", 5)] or "").strip() if "proveedor" in idx else ""
        estado = str(row[idx.get("estado", 6)] or "pendiente").strip().lower() if "estado" in idx else "pendiente"
        categoria = str(row[idx.get("categoria", 7)] or "").strip() if "categoria" in idx else ""
        notas = str(row[idx.get("notas", 8)] or "").strip() if "notas" in idx else ""

        row_errors = []
        if not fecha: row_errors.append("fecha invalida")
        if not descripcion: row_errors.append("descripcion vacia")
        if monto is None or monto <= 0: row_errors.append("monto invalido")
        if not empresa_str: row_errors.append("empresa vacia")
        if not banco_str: row_errors.append("banco vacio")

        empresa_id = maestros["empresas"].get(empresa_str.lower())
        banco_id = maestros["bancos"].get(banco_str.lower())
        if empresa_str and empresa_id is None:
            empresas_faltantes.add(empresa_str)
            row_errors.append(f"empresa '{empresa_str}' no existe")
        if banco_str and banco_id is None:
            bancos_faltantes.add(banco_str)
            row_errors.append(f"banco '{banco_str}' no existe")
        proveedor_id = maestros["proveedores"].get(proveedor_str.lower()) if proveedor_str else None

        if estado not in ("pendiente", "en_curso", "pagado", "cancelado", "rechazado"):
            warnings.append(f"fila {ri}: estado '{estado}' invalido, usa 'pendiente'")
            estado = "pendiente"

        items.append({
            "row": ri,
            "fecha_pago": fecha, "descripcion": descripcion, "monto": monto,
            "empresa_id": empresa_id, "empresa_str": empresa_str,
            "banco_id": banco_id, "banco_str": banco_str,
            "proveedor_id": proveedor_id, "proveedor_str": proveedor_str,
            "estado": estado, "categoria": categoria or None, "notas": notas or None,
            "errors": row_errors,
        })

    if empresas_faltantes:
        warnings.append(f"Empresas faltantes (cargalas primero): {', '.join(sorted(empresas_faltantes))}")
    if bancos_faltantes:
        warnings.append(f"Bancos faltantes (cargalos primero): {', '.join(sorted(bancos_faltantes))}")

    return {"items": items, "errors": errors, "warnings": warnings, "total": len(items)}


def aplicar_erogaciones(items: list[dict]) -> dict:
    """Inserta erogaciones validas (sin errors). Detecta duplicados por (empresa, monto ±5%, fecha ±2d)."""
    insertadas = 0
    salteadas_duplicado = 0
    salteadas_error = 0
    with get_conn() as c, c.cursor() as cur:
        for item in items:
            if item.get("errors"):
                salteadas_error += 1
                continue
            # Detectar duplicado
            cur.execute(
                """
                SELECT id FROM public."erogaciones"
                WHERE empresa_id = %s AND banco_id = %s
                  AND ABS(monto - %s) / NULLIF(monto, 0) < 0.05
                  AND ABS(fecha_pago - %s::date) <= 2
                LIMIT 1
                """,
                (item["empresa_id"], item["banco_id"], item["monto"], item["fecha_pago"]),
            )
            if cur.fetchone():
                salteadas_duplicado += 1
                continue
            cur.execute(
                """
                INSERT INTO public."erogaciones"
                  (fecha_pago, descripcion, monto, empresa_id, banco_id, proveedor_id,
                   estado, categoria, notas)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                """,
                (item["fecha_pago"], item["descripcion"], item["monto"],
                 item["empresa_id"], item["banco_id"], item.get("proveedor_id"),
                 item["estado"], item.get("categoria"), item.get("notas")),
            )
            insertadas += 1
    return {"insertadas": insertadas, "salteadas_duplicado": salteadas_duplicado, "salteadas_error": salteadas_error}


# ============================================================
# Plantilla ingresos puntuales
# ============================================================

PLANTILLA_INGRESOS_HEADERS = ["fecha", "descripcion", "monto", "empresa", "banco", "categoria", "notas"]


def generar_plantilla_ingresos() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "INGRESOS_PUNTUALES"
    for i, h in enumerate(PLANTILLA_INGRESOS_HEADERS, 1):
        ws.cell(row=1, column=i, value=h)
        ws.column_dimensions[get_column_letter(i)].width = 18
    ws.cell(row=2, column=1, value="2026-05-25")
    ws.cell(row=2, column=2, value="Cobro cheque diferido")
    ws.cell(row=2, column=3, value=500000)
    ws.cell(row=2, column=4, value="FOX ELECTRONICS")
    ws.cell(row=2, column=5, value="SUPERVIELLE")
    ws.cell(row=2, column=6, value="cobro_cheque")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parsear_ingresos(file_bytes: bytes) -> dict:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    if ws is None:
        return {"items": [], "errors": ["Excel vacio"], "warnings": []}
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return {"items": [], "errors": ["Solo headers"], "warnings": []}

    maestros = _load_maestros()
    items: list[dict] = []
    warnings: list[str] = []
    empresas_faltantes: set[str] = set()

    headers = [str(h).strip().lower() if h else "" for h in rows[0]]
    idx = {h: i for i, h in enumerate(headers)}

    for ri, row in enumerate(rows[1:], 2):
        if not any(c is not None and c != "" for c in row):
            continue
        fecha = _parse_date(row[idx.get("fecha", 0)] if "fecha" in idx else None)
        descripcion = str(row[idx.get("descripcion", 1)] or "").strip() if "descripcion" in idx else ""
        monto = _parse_number(row[idx.get("monto", 2)] if "monto" in idx else None)
        empresa_str = str(row[idx.get("empresa", 3)] or "").strip() if "empresa" in idx else ""
        banco_str = str(row[idx.get("banco", 4)] or "").strip() if "banco" in idx else ""
        categoria = str(row[idx.get("categoria", 5)] or "").strip() if "categoria" in idx else ""
        notas = str(row[idx.get("notas", 6)] or "").strip() if "notas" in idx else ""

        row_errors = []
        if not fecha: row_errors.append("fecha invalida")
        if not descripcion: row_errors.append("descripcion vacia")
        if monto is None or monto <= 0: row_errors.append("monto invalido")
        if not empresa_str: row_errors.append("empresa vacia")

        empresa_id = maestros["empresas"].get(empresa_str.lower())
        banco_id = maestros["bancos"].get(banco_str.lower()) if banco_str else None
        if empresa_str and empresa_id is None:
            empresas_faltantes.add(empresa_str)
            row_errors.append(f"empresa '{empresa_str}' no existe")
        if banco_str and banco_id is None:
            warnings.append(f"fila {ri}: banco '{banco_str}' no existe, se importa sin banco")

        items.append({
            "row": ri,
            "fecha": fecha, "descripcion": descripcion, "monto": monto,
            "empresa_id": empresa_id, "empresa_str": empresa_str,
            "banco_id": banco_id, "banco_str": banco_str,
            "categoria": categoria or None, "notas": notas or None,
            "errors": row_errors,
        })

    if empresas_faltantes:
        warnings.append(f"Empresas faltantes: {', '.join(sorted(empresas_faltantes))}")

    return {"items": items, "warnings": warnings, "total": len(items)}


def aplicar_ingresos(items: list[dict]) -> dict:
    insertadas = 0
    salteadas_error = 0
    with get_conn() as c, c.cursor() as cur:
        for item in items:
            if item.get("errors"):
                salteadas_error += 1
                continue
            cur.execute(
                """
                INSERT INTO public."ingresos_puntuales"
                  (fecha, descripcion, monto, empresa_id, banco_id, categoria, notas, origen)
                VALUES (%s, %s, %s, %s, %s, %s, %s, 'excel')
                """,
                (item["fecha"], item["descripcion"], item["monto"], item["empresa_id"],
                 item.get("banco_id"), item.get("categoria"), item.get("notas")),
            )
            insertadas += 1
    return {"insertadas": insertadas, "salteadas_error": salteadas_error}


# ============================================================
# Plantilla facturación
# ============================================================

PLANTILLA_FACTURACION_HEADERS = ["fecha", "unidad_negocio", "empresa", "monto", "es_real", "es_evento_puntual"]


def generar_plantilla_facturacion() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "FACTURACION"
    for i, h in enumerate(PLANTILLA_FACTURACION_HEADERS, 1):
        ws.cell(row=1, column=i, value=h)
        ws.column_dimensions[get_column_letter(i)].width = 18
    ws.cell(row=2, column=1, value="2026-05-25")
    ws.cell(row=2, column=2, value="Unistore Mayorista")
    ws.cell(row=2, column=3, value="UNISTORE")
    ws.cell(row=2, column=4, value=850000)
    ws.cell(row=2, column=5, value="si")
    ws.cell(row=2, column=6, value="no")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parsear_facturacion(file_bytes: bytes) -> dict:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    if ws is None:
        return {"items": [], "errors": ["Excel vacio"], "warnings": []}
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return {"items": [], "errors": ["Solo headers"], "warnings": []}

    maestros = _load_maestros()
    items: list[dict] = []
    warnings: list[str] = []
    unidades_faltantes: set[str] = set()

    headers = [str(h).strip().lower() if h else "" for h in rows[0]]
    idx = {h: i for i, h in enumerate(headers)}

    def truthy(v) -> bool:
        if isinstance(v, bool): return v
        s = str(v).strip().lower() if v else ""
        return s in ("si", "sí", "true", "1", "yes", "y", "x")

    for ri, row in enumerate(rows[1:], 2):
        if not any(c is not None and c != "" for c in row):
            continue
        fecha = _parse_date(row[idx.get("fecha", 0)] if "fecha" in idx else None)
        unidad_str = str(row[idx.get("unidad_negocio", 1)] or "").strip() if "unidad_negocio" in idx else ""
        empresa_str = str(row[idx.get("empresa", 2)] or "").strip() if "empresa" in idx else ""
        monto = _parse_number(row[idx.get("monto", 3)] if "monto" in idx else None)
        es_real = truthy(row[idx.get("es_real", 4)] if "es_real" in idx else True)
        es_evento_puntual = truthy(row[idx.get("es_evento_puntual", 5)] if "es_evento_puntual" in idx else False)

        row_errors = []
        if not fecha: row_errors.append("fecha invalida")
        if monto is None or monto < 0: row_errors.append("monto invalido")
        if not unidad_str: row_errors.append("unidad_negocio vacia")

        unidad_id = maestros["unidades_negocio"].get(unidad_str.lower())
        empresa_id = maestros["empresas"].get(empresa_str.lower()) if empresa_str else None
        if unidad_str and unidad_id is None:
            unidades_faltantes.add(unidad_str)
            row_errors.append(f"unidad '{unidad_str}' no existe")

        items.append({
            "row": ri,
            "fecha": fecha, "unidad_negocio_id": unidad_id, "unidad_str": unidad_str,
            "empresa_id": empresa_id, "empresa_str": empresa_str, "monto": monto,
            "es_real": es_real, "es_evento_puntual": es_evento_puntual,
            "errors": row_errors,
        })

    if unidades_faltantes:
        warnings.append(f"Unidades faltantes: {', '.join(sorted(unidades_faltantes))}")

    return {"items": items, "warnings": warnings, "total": len(items)}


def aplicar_facturacion(items: list[dict]) -> dict:
    insertadas = 0
    salteadas_error = 0
    with get_conn() as c, c.cursor() as cur:
        for item in items:
            if item.get("errors"):
                salteadas_error += 1
                continue
            cur.execute(
                """
                INSERT INTO public."facturacion_diaria"
                  (fecha, unidad_negocio_id, empresa_id, monto, es_real, es_evento_puntual, origen)
                VALUES (%s, %s, %s, %s, %s, %s, 'excel')
                """,
                (item["fecha"], item["unidad_negocio_id"], item.get("empresa_id"),
                 item["monto"], item["es_real"], item["es_evento_puntual"]),
            )
            insertadas += 1
    return {"insertadas": insertadas, "salteadas_error": salteadas_error}
