"""
Importer del Excel "VALOR PRODUCTO.xlsx" hacia las tablas cost_lote / cost_item.

Sheet esperado: "VALOR COMPRA Y PESO" (51 cols, ~3000 filas).
Replace-on-import por (Lote): si ya existe, se borra y reinserta.

Mapeo de columnas:
  Lote          -> cost_lote.lote
  Proveedor     -> cost_lote.proveedor
  Fecha Ingreso -> cost_lote.fecha_ingreso
  ORIGEN        -> cost_lote.origen
  Envio         -> cost_lote.envio
  MONEDA VAL    -> cost_lote.moneda
  SKU2          -> cost_item.sku
  PRODUCTO      -> cost_item.producto
  Categoria     -> cost_item.categoria
  Sub-Categoria -> cost_item.sub_categoria
  NCM           -> cost_item.ncm
  Cantidad      -> cost_item.cantidad
  valor maximo  -> cost_item.valor_max_usd
  valor minimo  -> cost_item.valor_min_usd
  Costo Total S/ IVA Max (USD) -> cost_item.costo_total_sin_iva_usd
  Costo con IVA -> cost_item.costo_con_iva_usd
  Precio        -> cost_item.precio_ars
  Rentabilidad  -> cost_item.rentabilidad_ars
  % Rentab      -> cost_item.pct_rentabilidad
  Alto/Largo/Ancho (m) -> alto_m / largo_m / ancho_m
  Peso grueso Unitario -> peso_kg
  CBM (un)      -> cbm_un
  (resto)       -> raw_payload (jsonb)
"""
from __future__ import annotations

import datetime as dt
import io
import logging
from typing import IO

import openpyxl

from app.db import costs_db

log = logging.getLogger(__name__)

SHEET_NAME = "VALOR COMPRA Y PESO"

# Nombres exactos de columnas en el header del Excel (case-sensitive con strip).
COLS_LOTE = {
    "lote": "Lote",
    "proveedor": "Proveedor",
    "fecha_ingreso": "Fecha Ingreso",
    "origen": "ORIGEN",
    "envio": "Envio",
    "moneda": "MONEDA VAL",
}

COLS_ITEM = {
    "sku": "SKU2",
    "producto": "PRODUCTO",
    "categoria": "Categoria",
    "sub_categoria": "Sub-Categoria",
    "ncm": "NCM",
    "cantidad": "Cantidad",
    "valor_max_usd": "valor maximo (todo manual salvo combos uqe hay que hacer formula)",
    "valor_min_usd": "valor minimo",
    # Costos per-unit landed (USD)
    "costo_unit_usd_max": "Costo Total S/ IVA Max (USD)",
    "costo_unit_usd_min": "Costo Total S/ IVA Min (USD)2",
    # Costos per-unit ARS (la planilla calcula con TC al importar)
    "costo_unit_ars": "Costo Total S/ IVA ",  # nota: trailing space en el Excel
    "costo_con_iva_unit_ars": "Costo con IVA",
    "precio_ars": "Precio ",
    "rentabilidad_ars": "Rentabilidad",
    "pct_rentabilidad": "% Rentab",
    "rent_neta_lote_ars": "Rent, Neta Lote",
    "facturacion_ars": "Facturacion",
    "alto_m": "Alto (m)",
    "largo_m": "Largo (m)",
    "ancho_m": "Ancho (m)",
    "peso_kg": "Peso grueso Unitario",
    "cbm_un": "CBM (un)",
}


def _to_float(v) -> float | None:
    if v is None or v == "" or v == "-":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _to_int(v) -> int | None:
    f = _to_float(v)
    return int(f) if f is not None else None


def _to_str(v) -> str | None:
    if v is None or v == "":
        return None
    s = str(v).strip()
    return s if s and s != "-" else None


def _to_date_str(v) -> str | None:
    """Devuelve YYYY-MM-DD o None."""
    if v is None or v == "" or v == "-":
        return None
    if isinstance(v, dt.datetime):
        return v.date().isoformat()
    if isinstance(v, dt.date):
        return v.isoformat()
    s = str(v).strip()
    # Intentos varios formatos
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return dt.datetime.strptime(s[:10], fmt).date().isoformat()
        except ValueError:
            continue
    return None


def parse_excel(file_bytes: bytes | IO[bytes]) -> dict:
    """Parsea el .xlsx y devuelve estructura agrupada por lote.

    Returns:
      {
        "rows_total": int,
        "rows_skipped": int,
        "lotes": {
            "<lote_name>": {
                "lote_meta": {...},
                "items": [{...}, ...]
            }
        }
      }
    """
    if isinstance(file_bytes, bytes):
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    else:
        wb = openpyxl.load_workbook(file_bytes, data_only=True, read_only=True)

    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f'El archivo no tiene la hoja "{SHEET_NAME}". Hojas disponibles: {wb.sheetnames}')

    ws = wb[SHEET_NAME]
    rows_iter = ws.iter_rows(values_only=True)

    # Header: primera fila no vacia
    header_row = next(rows_iter, None)
    if not header_row:
        raise ValueError("Archivo vacio")

    # Mapear nombre de columna -> indice
    header = [str(c).strip() if c is not None else "" for c in header_row]
    col_idx: dict[str, int] = {}
    for h, name in {**COLS_LOTE, **COLS_ITEM}.items():
        try:
            col_idx[h] = header.index(name)
        except ValueError:
            log.warning(f'Columna "{name}" no encontrada en el header. Sera tratada como nula.')
            col_idx[h] = -1

    lotes: dict[str, dict] = {}
    rows_total = 0
    rows_skipped = 0

    def cell(row, key: str):
        i = col_idx.get(key, -1)
        return row[i] if 0 <= i < len(row) else None

    for row in rows_iter:
        if row is None or all(c is None or c == "" for c in row):
            continue
        rows_total += 1

        lote_name = _to_str(cell(row, "lote"))
        sku = _to_str(cell(row, "sku"))
        if not lote_name or not sku:
            rows_skipped += 1
            continue

        # Meta del lote (la primera vez)
        if lote_name not in lotes:
            lotes[lote_name] = {
                "lote_meta": {
                    "lote": lote_name,
                    "proveedor": _to_str(cell(row, "proveedor")),
                    "fecha_ingreso": _to_date_str(cell(row, "fecha_ingreso")),
                    "origen": _to_str(cell(row, "origen")),
                    "envio": _to_str(cell(row, "envio")),
                    "moneda": _to_str(cell(row, "moneda")),
                },
                "items": [],
            }

        # Item
        # raw_payload: todo el row con sus headers para no perder columnas extras
        raw = {}
        for i, h in enumerate(header):
            if not h:
                continue
            v = row[i] if i < len(row) else None
            if isinstance(v, (dt.date, dt.datetime)):
                v = v.isoformat()
            raw[h] = v

        costo_unit_usd_max = _to_float(cell(row, "costo_unit_usd_max"))
        costo_unit_usd_min = _to_float(cell(row, "costo_unit_usd_min"))
        costo_unit_ars = _to_float(cell(row, "costo_unit_ars"))
        costo_con_iva_unit_ars = _to_float(cell(row, "costo_con_iva_unit_ars"))

        item = {
            "sku": sku,
            "producto": _to_str(cell(row, "producto")),
            "categoria": _to_str(cell(row, "categoria")),
            "sub_categoria": _to_str(cell(row, "sub_categoria")),
            "ncm": _to_str(cell(row, "ncm")),
            "cantidad": _to_int(cell(row, "cantidad")),
            "valor_max_usd": _to_float(cell(row, "valor_max_usd")),
            "valor_min_usd": _to_float(cell(row, "valor_min_usd")),
            # Costo per-unit corregido (campos nuevos)
            "costo_unit_usd_max": costo_unit_usd_max,
            "costo_unit_usd_min": costo_unit_usd_min,
            "costo_unit_ars": costo_unit_ars,
            "costo_con_iva_unit_ars": costo_con_iva_unit_ars,
            # Backwards-compat de campos antiguos
            "costo_total_sin_iva_usd": costo_unit_usd_max,
            "costo_con_iva_usd": None,
            "precio_ars": _to_float(cell(row, "precio_ars")),
            "rentabilidad_ars": _to_float(cell(row, "rentabilidad_ars")),
            "pct_rentabilidad": _to_float(cell(row, "pct_rentabilidad")),
            "rent_neta_lote_ars": _to_float(cell(row, "rent_neta_lote_ars")),
            "facturacion_ars": _to_float(cell(row, "facturacion_ars")),
            "alto_m": _to_float(cell(row, "alto_m")),
            "largo_m": _to_float(cell(row, "largo_m")),
            "ancho_m": _to_float(cell(row, "ancho_m")),
            "peso_kg": _to_float(cell(row, "peso_kg")),
            "cbm_un": _to_float(cell(row, "cbm_un")),
            "raw_payload": raw,
        }
        lotes[lote_name]["items"].append(item)

    return {
        "rows_total": rows_total,
        "rows_skipped": rows_skipped,
        "lotes": lotes,
    }


def import_excel(file_bytes: bytes | IO[bytes], imported_by: str, source_file: str = "VALOR PRODUCTO.xlsx") -> dict:
    """Parsea + persiste todos los lotes encontrados en cost_lote/cost_item.

    Devuelve resumen para mostrarle al user.
    """
    parsed = parse_excel(file_bytes)
    summary = {
        "source_file": source_file,
        "rows_total": parsed["rows_total"],
        "rows_skipped": parsed["rows_skipped"],
        "lotes_processed": 0,
        "lotes_replaced": 0,
        "items_imported": 0,
        "lote_results": [],
    }

    for lote_name, payload in parsed["lotes"].items():
        meta = payload["lote_meta"]
        items = payload["items"]
        try:
            res = costs_db.upsert_lote(
                lote=meta["lote"],
                proveedor=meta.get("proveedor"),
                fecha_ingreso=meta.get("fecha_ingreso"),
                origen=meta.get("origen"),
                envio=meta.get("envio"),
                moneda=meta.get("moneda"),
                source_file=source_file,
                imported_by=imported_by,
                items=items,
            )
            summary["lotes_processed"] += 1
            if res.get("replaced"):
                summary["lotes_replaced"] += 1
            summary["items_imported"] += res.get("items_count", 0)
            summary["lote_results"].append({
                "lote": lote_name,
                "items_count": res.get("items_count", 0),
                "replaced": res.get("replaced", False),
            })
        except Exception as e:
            log.warning(f"Error importing lote {lote_name}: {e}")
            summary["lote_results"].append({
                "lote": lote_name,
                "error": str(e),
            })

    return summary
